# ------------------------------------------------------------------
#  Django
# ------------------------------------------------------------------
from django.contrib               import admin
from django.contrib.admin         import SimpleListFilter
from django.contrib.admin.models  import LogEntry
from django.contrib.contenttypes.models import ContentType
from django.db.models             import OuterRef, Subquery, F, CharField
from django.db.models.functions   import Cast
from django.http                  import HttpResponse
from rangefilter.filters import DateRangeFilter
# ------------------------------------------------------------------
#  Import-Export
# ------------------------------------------------------------------
from import_export                import resources
from import_export.admin          import ExportActionModelAdmin

# ------------------------------------------------------------------
#  OpenPyXL
# ------------------------------------------------------------------
from openpyxl                     import Workbook
from openpyxl.styles              import Font, PatternFill, Alignment
from openpyxl.utils               import get_column_letter

# ------------------------------------------------------------------
#  Modelos locales
# ------------------------------------------------------------------
from .models import (
    CentroEmpadronamiento, Coordinador, Estacion, Item,
    Llave, MovimientosEstacion, Operador, RegistroDespliegue,
    ReporteDiario, Ruta, UbicacionesOperador, Megacentro,
    ReporteDiarioLogistico
)      
# ------------------------------------------------------------------
#  Historico
# ------------------------------------------------------------------
from simple_history.admin import SimpleHistoryAdmin

def _primer_que(modelo, obj_id):
    log = LogEntry.objects.filter(
        content_type=ContentType.objects.get_for_model(modelo),
        object_id=str(obj_id)
    ).order_by('action_time').first()
    return log.user.username if log else '-'

def _ultimo_que(modelo, obj_id):
    log = LogEntry.objects.filter(
        content_type=ContentType.objects.get_for_model(modelo),
        object_id=str(obj_id)
    ).order_by('-action_time').first()
    return log.user.username if log else '-'

def exportar_a_excel_reportes_diarios(modeladmin, request, queryset):
    """
    Opcion para exportar los registros seleccionados a excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Registros Diarios'

    # --- cabecera ---
    columnas = [
        "Departamento",
        "Provincia",
        "Municipio",
        "Punto de Empadronamiento",
        "Direccion",
        "Fecha",
        "Estación",
        "Cant. de registros C",
        "Cant. de registros R",
        "Tipo de estación",
        "Área",
        "Operador",
        "Observación"
    ]
    #columnas = [f.verbose_name for f in ReporteDiario._meta.fields]
    ws.append(columnas)

    for col_num, col_name in enumerate(columnas, 1):
        col_letter = get_column_letter(col_num)

        #color de fondo
        ws[f"{col_letter}1"].fill = PatternFill(
            start_color='ADD8E6', #Azul claro,
            end_color='ADD8E6',
            fill_type="solid"
        )

        #Fuente negrita
        ws[f"{col_letter}1"].font= Font(bold=True)

        #Centrado
        ws[f"{col_letter}1"].alignment= Alignment(horizontal='center')

        ws.column_dimensions[col_letter].width = max(15 , len(col_name) + 3)

    # --- filas ---
    for r in queryset:
        # fila = []
        # for f in ReporteDiario._meta.fields:
        #     valor = getattr(obj, f.name)   

        #     if f.many_to_one:
        #         valor = str(valor)
            
        #     from datetime import datetime
        #     if isinstance(valor, datetime):
        #         valor = valor.strftime("%Y-%m-%d")
            
        #     fila.append(valor)
        # ws.append(fila)
        ws.append([
            r.centro_empadronamiento.departamento if r.centro_empadronamiento else "",
            r.centro_empadronamiento.provincia if r.centro_empadronamiento else "",
            r.centro_empadronamiento.municipio if r.centro_empadronamiento else "",
            r.centro_empadronamiento.punto_de_empadronamiento if r.centro_empadronamiento else "",
            r.centro_empadronamiento.direccion if r.centro_empadronamiento else "",
            r.fecha_reporte.strftime("%Y-%m-%d"),
            r.estacion.nro_estacion if r.estacion else "",
            r.registro_c,
            r.registro_r,
            r.estacion.tipo_estacion if r.estacion else "",
            r.operador.tipo_operador if r.operador else "",
            f"{r.operador.nombre} {r.operador.apellido_paterno} {r.operador.apellido_materno}",
            r.observaciones
        ])
    
    response = HttpResponse(
        content_type= 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_diario.xlsx"'
    wb.save(response)
    return response

exportar_a_excel_reportes_diarios.short_description = "Exportar registros diarios a Excel"

class registroDiarioResource(resources.ModelResource):
    class Meta:
        # modelo de donde se exportan los datos
        model = ReporteDiario
        # exportar todos los modelos
        fields = '__all__'

class EstacionAdmin(admin.ModelAdmin):
    def get_queryset(self, request):
        ct = ContentType.objects.get_for_model(Estacion)
        ultimo_qs = LogEntry.objects.filter(
            content_type=ct,
            object_id=OuterRef('pk')
        ).order_by('-action_time').values('user__username')[:1]

        return super().get_queryset(request).annotate(
            _ultimo_user=Subquery(ultimo_qs),
        )

class UltimoUsuarioFilter(SimpleListFilter):
    title          = 'último usuario'
    parameter_name = 'ultimo_user'

    def lookups(self, request, model_admin):
        # usuarios que aparecen en LogEntry sobre Estacion
        ct = ContentType.objects.get_for_model(model_admin.model)
        users = (LogEntry.objects
                 .filter(content_type=ct)
                 .values_list('user__username', flat=True)
                 .distinct()
                 .order_by('user__username'))
        return [(u, u) for u in users]

    def queryset(self, request, queryset):
        if not self.value():
            return queryset
        ct = ContentType.objects.get_for_model(queryset.model)
        # TEXT = TEXT
        ultimo_qs = (LogEntry.objects
                     .filter(content_type=ct,
                             object_id=Cast(OuterRef('pk'), output_field=CharField()))  # ← cast
                     .order_by('-action_time')
                     .values('user__username')[:1])
        return queryset.annotate(_ultimo_user=Subquery(ultimo_qs)).filter(
            _ultimo_user=self.value()
        )

def exportar_a_excel(modeladmin, request, queryset):
    """
    Acción personalizada: Exportar registros seleccionados a Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Estaciones"

    # --- cabecera ---
    campos = [f.verbose_name for f in Estacion._meta.fields]
    ws.append(campos)

    # --- filas ---
    for obj in queryset:
        ws.append([getattr(obj, f.name) for f in Estacion._meta.fields])

    # --- respuesta HTTP con el archivo ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="estaciones_seleccionadas.xlsx"'
    wb.save(response)
    return response

exportar_a_excel.short_description = "Exportar a Excel"

class EstacionResource(resources.ModelResource):
    """Resource para elegir qué columnas exportar"""
    class Meta:
        model = Estacion
        # exportar TODOS los campos (o lista los que quieras)
        fields = '__all__'
        # exclude = ['observacion']   # si quieres omitir alguno


# ---------------- Estacion  ----------------
@admin.register(Estacion)
class EstacionAdmin(SimpleHistoryAdmin):
    list_display = [
        'codigo_equipo',
        'llave',
        'fase',
        'ultimo_por',
    ]
    search_fields = [
        'codigo_equipo',
        'nro_estacion',
        'fase',
        'estado_computadora',
        'llave__nro_estacion',   # búsqueda por número de llave
    ]
    list_filter  = [UltimoUsuarioFilter, 'fase'] 
    def primer_por(self, obj):
        return _primer_que(Estacion, obj.pk)
    primer_por.short_description = 'Primero por'

    def ultimo_por(self, obj):
        ct = ContentType.objects.get_for_model(obj)
        log = (LogEntry.objects
            .filter(content_type=ct,
                    object_id=str(obj.pk))        # str() basta aquí
            .order_by('-action_time')
            .first())
        return log.user.username if log else '-'
    ultimo_por.short_description = 'Último por'



    actions = [exportar_a_excel]

    fieldsets = (
        ('Datos generales', {
            'fields': ('codigo_equipo', 'nro_estacion', 'fase', 'tipo_estacion', 'modelo', 'llave')
        }),
        ('Hardware', {
            'description': 'Estado del hardware y sus observaciones',
            'fields': (
                ('estado_computadora', 'obs_computadora'),
                ('estado_monitor', 'obs_monitor'),
                ('estado_pad_firmas', 'obs_pad_firmas'),
                ('estado_decadactilar', 'obs_decadactilar'),
                ('estado_hub_usb', 'obs_hub_usb'),
                ('estado_estabilizador_energia', 'obs_estabilizador_energia'),
                ('estado_pila_madre', 'obs_pila_madre'),
                ('estado_memorias_ram', 'obs_memorias_ram'),
                ('estado_disco_duro', 'obs_disco_duro'),
                ('estado_teclado', 'obs_teclado'),
                ('estado_regulador_voltaje', 'obs_regulador_voltaje'),
                ('estado_escaner', 'modelo_escaner', 'obs_escaner'),
                ('estado_impresora', 'modelo_impresora', 'obs_impresora'),
                ('estado_camara', 'modelo_camara', 'obs_camara'),
            )
        }),
        ('Accesorios', {
            'fields': (
                ('cable_extensor', 'obs_cable_extensor'),
                ('tripode',        'obs_tripode'),
                ('banner',         'obs_banner'),
                ('adaptador_3a2',  'obs_adaptador_3a2'),
                'monitor_pc',
                'testeo_pila'
            )
        }),
        ('Otros', {
            'fields': ('asignada', 'observacion')
        })
    )


# ---------------- Llave ----------------
@admin.register(Llave)
class LlaveAdmin(admin.ModelAdmin):
    list_display = ['nro_estacion', 'contador_r', 'contador_c']
    search_fields = ['nro_estacion']
    list_filter = ['nro_estacion']

# ---------------- MovimientosEstacion ----------------
@admin.register(MovimientosEstacion)
class MovimientosEstacionAdmin(admin.ModelAdmin):
    list_display = ['estacion', 'fecha_movimiento']
    search_fields = ['estacion__codigo_equipo']
    list_filter = ['fecha_movimiento']

# ---------------- Coordinador ----------------
@admin.register(Coordinador)
class CoordinadorAdmin(admin.ModelAdmin):
    list_display = ['user', 'estado',  'correo', 'celular']
    search_fields = ['user__username', 'nombre', 'apellido_paterno', 'correo', 'celular']
    list_filter = ['estado']

#------------------ Reporte diario Logistico
@admin.register(ReporteDiarioLogistico)
class ReporteDiarioLogisticoAdmin(admin.ModelAdmin):
    list_display = ['nro_estacion', 'cantidad_de_formulario_faltantes', 'r_inicial', 'r_final', 'c_inicial', 'c_final', 'fecha_reporte']
    search_fields = ['nro_estacion']
    list_filter = ['nro_estacion', 'operador', 'fecha_reporte']
    date_hierarchy = 'fecha_reporte'

# ---------------- Operador ----------------
@admin.register(Operador)
class OperadorAdmin(admin.ModelAdmin):
    list_display = ['user__username', 'nombre','apellido_paterno', 'apellido_materno', 'carnet','tipo_operador', 'estado', 'coordinador', 'estacion']
    search_fields = ['user__username', 'nombre', 'apellido_paterno', 'correo', 'celular', 'carnet']
    list_filter = ['estado', 'tipo_operador', 'coordinador', 'estacion']

# ---------------- ReporteDiario ----------------
@admin.register(ReporteDiario)
class ReporteDiarioAdmin(admin.ModelAdmin):
    #resource_classes = [registroDiarioResource]
    list_display = ['estacion', 'fecha_reporte', 'operador__tipo_operador', 'operador', 'registro_c', 'registro_r', 'estado']
    search_fields = ['operador__user__username', 'estacion__codigo_equipo', 'estacion__nro_estacion']
    list_filter = [
        ('fecha_reporte', DateRangeFilter),
        'estado', 
        'operador__tipo_operador']
    date_hierarchy = 'fecha_reporte'
    actions = [exportar_a_excel_reportes_diarios]

# ---------------- RegistroDespliegue ----------------
@admin.register(RegistroDespliegue)
class RegistroDespliegueAdmin(admin.ModelAdmin):
    list_display = ['operador', 'centro_empadronamiento', 'descripcion_reporte', 'fecha_hora', 'sincronizar']
    search_fields = ['operador__user__username', 'destino']
    list_filter = ['descripcion_reporte', 'sincronizar', 'fecha_hora']

# ---------------- Item ----------------
@admin.register(Item)
class ItemAdmin(admin.ModelAdmin):
    list_display = ['codigo_item', 'serie_item', 'tipo', 'asignado_operador']
    search_fields = ['codigo_item', 'serie_item']
    list_filter = ['tipo', 'asignado_operador']

#--------------- Centro de empadronamiento ---------------------
@admin.register(CentroEmpadronamiento)
class CentroEmpadronamientoAdmin(admin.ModelAdmin):
    list_display = [f.name for f in CentroEmpadronamiento._meta.fields]
    search_fields = ['nombre', 'provincia']
    list_filter = ['provincia']

#--------------- UbicacionesOperador ---------------------
@admin.register(UbicacionesOperador)
class UbicacionesOperadorAdmin(admin.ModelAdmin):
    list_display = [f.name for f in UbicacionesOperador._meta.fields]
    search_fields = ['operador', 'operador__carnet']
    list_filter = ['fecha', 'operador']


#--------------- Megacentro ---------------------
@admin.register(Megacentro)
class MegacentroAdmin(admin.ModelAdmin):
    list_display = ['nombre', 'descripcion', 'fecha_creacion', 'coordinador', 'soporte', 'logistico', 'asistente_megacentro']
    search_fields = ['nombre']
    list_filter = ['fecha_creacion']

# ---------------- Ruta ---------------------
@admin.register(Ruta)
class RutaAdmin(admin.ModelAdmin):
    list_display = ['nombre', 'descripcion', 'fecha_creacion', 'coordinador', 'soporte', 'logistico']
    search_fields = ['nombre']
    list_filter = ['fecha_creacion']
